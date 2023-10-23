import sys
import csv
import re
import getopt
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.styles import Alignment
from openpyxl.comments import Comment

def process(testfile, summaryfile, logfile):

    # Load the test-definition file
    print( 'Loading automation test file: {0}...'.format(testfile))
    test_definitions = {}
    tf = open(testfile, 'r')
    readfile = csv.reader(tf)
    for i,row in enumerate(readfile):
        if (i == 0): # Skip first line because its a header line
            header = row
            continue
        if len(row) == 0: # Skip empty line
            continue
        test_definitions[row[0]] = row[1:]
    tf.close()
    del tf

    # Load the summary file
    print( 'Loading automation summary file: {0}...'.format(summaryfile))
    summary = {}
    sf = open(summaryfile, 'r')
    readfile = csv.reader(sf)
    for i,row in enumerate(readfile):
        if (i == 0): # Skip first line because its a header line
            header = row
            continue
        summary[row[0]] = row[1:]
    sf.close()
    del sf

    # Load the log file
    print( 'Loading automation log file: {0}...'.format(logfile))
    key = 'common'
    log = { key: [] }
    af = open(logfile, 'r')
    for i,row in enumerate(af.readlines()):

        row = row.replace('\n', '')

        m = re.match( r"^.*Test ([^\s]+) (RE)?START", row)
        if m is not None:
            key = m.groups()[0]
            log[key] = []

        log[key].append(row)

        m = re.match( r"^.*Test with DSLAM value " + key + " ((Fails)|(Succeeded))", row)
        if m is not None:
            key = 'common'

    af.close()
    del af

    print( 'Number of lines in log file: {0}'.format(i+1) )
    j = 0
    for x in log.values():
        j += len(x)
    print( 'Number of lines in loaded log dictionary: {0}'. format(j) )
    del log['common']

    # Check that the set of keys in summary and log files are same
    diff = set(log.keys()).union(set(summary.keys())).difference(
                    set(log.keys()).intersection(set(summary.keys())))
    if len(diff):
        print( 'Following tests are missing from one of the summary or log files: {0}'.format(", ".join(diff)))
        #sys.exit(0)

    print( "{0} test-id's in test-definition file: {1}".format(len(test_definitions), list(test_definitions.keys())))
    print( "{0} test-id's in summary file: {1}".format(len(summary), list(summary.keys())))
    print( "{0} test-id's in log file: {1}".format(len(log), list(log.keys())))

    xlsx_filename = "report.xlsx"
    xlsx = openpyxl.Workbook()

    sheet = xlsx.active

    columns = [ 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
                'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM',
                'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ' ]

    #==============================================================
    # SUMMARY
    #==============================================================

    ft = Font( bold=True )

    fence_border = Border(left=Side(style='double'), right=Side(style='double'))
    lower_border = Border(bottom=Side(style='medium'), right=Side(style='thin'))
    left_border = Border(left=Side(style='medium'))
    top_border = Border(top=Side(style='medium'), right=Side(style='thin'), bottom=Side(style='thin'))
    top_border_only = Border(top=Side(style='medium'))
    top_and_left_border = Border(top=Side(style='medium'), left=Side(style='medium'), right=Side(style='thin'), bottom=Side(style='thin'))
    lower_and_left_border = Border(bottom=Side(style='medium'), left=Side(style='medium'), right=Side(style='thin'))

    gray_accent3_lighter_80percent = PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type="solid")
    orange_accent2_lighter_80percent = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type="solid")
    gold_accent4_lighter_80percent = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type="solid")
    white_background1_darker_15percent = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type="solid")
    blue_accent1_lighter_80percent = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type="solid")
    green_accent6_lighter_80percent = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type="solid")

    # Add summary header row 2:
    # ------------------------
    for i,title in enumerate([ 'DSLAM','','TXIN', '', '', 'Link 0', '', 'Link 1', '', '', '',
                               '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                               '', '', '', '', '', '', '', '' ]):
        cellid = columns[i+1] + '2'
        sheet[cellid] = title
        sheet[cellid].alignment = Alignment(wrapText=False, horizontal='center', vertical='center')
        sheet[cellid].font = ft

        # Border:
        if i == 16:
            sheet[cellid].border = fence_border

        # Color:
        if i == 5:
            sheet[cellid].fill = gray_accent3_lighter_80percent
        elif i == 7:
            sheet[cellid].fill = orange_accent2_lighter_80percent
        elif i == 15:
            sheet[cellid].fill = white_background1_darker_15percent 
        elif i > 16:
            sheet[cellid].fill = gold_accent4_lighter_80percent 

    # merge cells H2 with G2. And also J2 with I2
    sheet.merge_cells('G2:H2')
    sheet.merge_cells('I2:J2')

    # Add summary header row 3:
    # ------------------------
    for i,title in enumerate([ '','','','Rate\nRatio\n(m)','Rate\nRatio\n(n)',
                               'Rate\n[kbit/s]', 'ActIntlDelay\n[ms]', 'Rate\n[kbit/s]', 'ActIntlDelay\n[ms]',
                               'TxDelay\nLine-0', 'TxDelay\nLine-1', 'HW-Shaper', 'SW-Shaper\nPercentage',
                               'SW-Shaper\nRate', 'US\nDiffIntlDelay\nCompens\n[ms]', 'US\niperf\ntraffic\nrate',
                               'Expected\nResult',
                               'Test ID',
                               'Expected\nThrougput', 'Actual\nThroughput', 'loss',
                               'Highest\nFill Level\n(0)', 'Highest\nFill Level\n(1)', 
                               'Line-0\nRate\nafter\nIperf', 'Line-1\nRate\nafter\nIperf',
                               'Time slab counters', '', '', '', '', '', '', '', '', '']):

        cellid = columns[i+1] + '3'
        sheet[cellid] = title
        sheet[cellid].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet[cellid].font = ft

        # Border:
        if i == 0:
            sheet[cellid].border = top_and_left_border
        elif i == 16:
            sheet[cellid].border = fence_border
        else:
            sheet[cellid].border = top_border

        # Color:
        if i == 14:
            sheet[cellid].fill = blue_accent1_lighter_80percent 
        elif i == 15:
            sheet[cellid].fill = white_background1_darker_15percent 

    # merge cells Z3 through AI3
    sheet.merge_cells('AA3:AJ3')

    # Add summary header row 4:
    # ------------------------
    for i,title in enumerate([ '','','','(0)','(1)', 'US Rate (0)', 'DelayUs(0)', 'US Rate (1)', 'DelayUs(1)', 
                               '[ms]', '[ms]', '[kbit/s]', '[%]', '[kbit/s]', '[ms]', '[kbit/s]',
                               '', '', '[Mbits/s]', '[Mbits/s]', '', '', '', '[Bits/s]', '[Bits/s]', 
                               '(0,.5]', '(.5,.75]', '(.75,1]', '(1,1.5]', '(1.5,2]',
                               '(2,2.5]', '(2.5,3]', '(3,5]', '(5,10]', '>10']):

        cellid = columns[i+1] + '4'
        sheet[cellid] = title
        sheet[cellid].alignment = Alignment(wrapText=False, horizontal='center', vertical='center')

        # Border:
        if i == 0:
            sheet[cellid].border = lower_and_left_border
        elif i == 16:
            sheet[cellid].border = fence_border
        else:
            sheet[cellid].border = lower_border

        # Color:
        if i >= 5 and i <= 11:
            sheet[cellid].fill = blue_accent1_lighter_80percent 
        elif i == 12 or i == 13:
            sheet[cellid].fill = green_accent6_lighter_80percent
        elif i == 14:
            sheet[cellid].fill = blue_accent1_lighter_80percent 
        elif i == 15:
            sheet[cellid].fill = white_background1_darker_15percent 
        elif i > 16:
            sheet[cellid].fill = gold_accent4_lighter_80percent 

    # Add Summary Header row 5:
    # ------------------------
    cellid = columns[1] + '5'
    sheet[cellid].border = left_border
    cellid = columns[17] + '5'
    sheet[cellid].border = fence_border

    # Add summary
    # -----------
    ft = Font( color='2F75B5' )
    summary_offset = { 'col': 18, 'row': 5 }
    col_to_test_defn_map = { 3:1, 4:2, 5:3, 6:4, 8:5, 10:6, 11:7, 13:8, 14:9 }
    for i, (test, values) in enumerate(summary.items()):

        for col in range(summary_offset['col']):
            cellid = columns[col] + str(summary_offset['row'] + i+1)
            if col in col_to_test_defn_map:
                sheet[cellid] = test_definitions[test][col_to_test_defn_map[col] - 1]
                sheet[cellid].alignment = Alignment(wrapText=False, horizontal='center')
            else:
                sheet[cellid] = ''
            if col == 1:
                sheet[cellid].border = left_border
            elif col == (summary_offset['col'] - 1):
                sheet[cellid].border = fence_border

        # Fill the test-description column
        cellid = columns[summary_offset['col']+0] + str(summary_offset['row'] + i+1)
        sheet[cellid] = test

        # Fill the test-result columns
        for j, value in enumerate(values):
            if j == 0 or j == 1:
                try:
                    value = str(round(float(value), 2))
                except:
                    pass

            cellid = columns[summary_offset['col']+j+1] + str(summary_offset['row'] + i+1)
            sheet[cellid] = value;
            sheet[cellid].alignment = Alignment(wrapText=False, horizontal='center')
            sheet[cellid].font = ft

    # Add summary trailer row 1:
    # ------------------------
    for col in range(35):
        cellid = columns[col+1] + str(summary_offset['row'] + i + 2)

        # Border:
        if col != (summary_offset['col'] - 2):
            sheet[cellid].border = top_border_only

    #==============================================================
    # LOG
    #==============================================================
    log_offset = { 'col': 1, 'row': int(re.match( r"[^:]+:[A-Z]+([0-9]+)", sheet.dimensions).groups()[0]) + 2 }
    logidx = 0
    pos = {'start': 0, 'end': 0}
    ft = Font( name='Courier New', size=11, bold=False, italic=False, underline='none', color='000000' )
    for key, value in log.items():

        if key == 'common':
            continue

        pos['start'] = log_offset['row'] + logidx + 1
        for line in value:
            logidx += 1
            cellid = columns[log_offset['col']+0] + str(log_offset['row'] + logidx)
            sheet[cellid] = line
            sheet[cellid].font = ft
        pos['end'] = log_offset['row'] + logidx

        # Add few more blank lines:
        for i in range(3):
            logidx += 1
            cellid = columns[log_offset['col']+0] + str(log_offset['row'] + logidx)
            sheet[cellid] = ""
            sheet[cellid].font = ft

        # Find the summary record:
        for si in range(summary_offset['row'], log_offset['row'] - 1):
            if key == sheet[columns[summary_offset['col']] + str(si)].value:
                sheet[columns[summary_offset['col']] + str(si)].comment = Comment( 'From row {0} to row {1}'.format(pos['start'],pos['end']),'MXL' )
                link = xlsx_filename + '#' + 'Sheet!B' + str(pos['start']) + ':P' + str(pos['end'])
                for iloc in range(len(summary[key])):
                    sheet[columns[summary_offset['col']+iloc] + str(si)].hyperlink = (link)

    #==============================================================

    # Job done. Save the file
    xlsx.save( xlsx_filename )
    print( "Generating report file: {0}".format(xlsx_filename))

#####################################################################
# MAIN
#####################################################################

def showhelp(progname):
    print("Usage: {0} options".format(progname))
    print("       {0} -h|--help".format(progname))
    print(" Options:")
    print("  -t|--test=<testcsvfilename>    Automation Test Definition file to load. If not specified, regular_test.csv is searched")
    print("  -s|--summary=<summaryfilename> Automation Summary file to load. If not specified, Summary.csv is searched")
    print("  -l|--log=<logfilename>         Automation Log file to load. If not specified, Automation.log is searched")
    print("  -h|--help                      print this message and quit")

def main(argv):

    testfile = "regular_test.csv"
    summaryfile = "Summary.csv"
    logfile = "Automation.log"

    try:
        opts, args = getopt.getopt(argv[1:], "hs:l:t:", ["help", "summary=", "log=", "test="])
    except getopt.GetoptError:
        showhelp(argv[0])
        sys.exit(0)

    for opt, arg in opts:
        if opt == '-t':
            testfile = arg
        if opt == '-s':
            summaryfile = arg
        elif opt == '-l':
            logfile = arg
        elif opt == '-h':
            showhelp(argv[0])
            sys.exit(0)

    process(testfile, summaryfile, logfile)

if __name__ == "__main__":
    main(sys.argv)
    print( 'Done')

